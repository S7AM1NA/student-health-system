"""
data_io.py - 数据导入导出工具 (Member C)
功能: 
1. Excel 导出用户健康数据 (睡眠、运动、饮食、身体指标)
2. Excel 导入食物数据库
3. JSON 导出用户数据备份
"""
import json
import csv
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def check_openpyxl():
    """检查 openpyxl 是否可用"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")


class ExcelExporter:
    """Excel 导出工具类"""
    
    @classmethod
    def _get_styles(cls):
        """获取 Excel 样式（仅在 openpyxl 可用时调用）"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        return {
            'header_font': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'header_alignment': Alignment(horizontal="center", vertical="center"),
            'thin_border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    @classmethod
    def export_user_health_data(cls, user, start_date=None, end_date=None):
        """
        导出用户的完整健康数据到 Excel
        
        参数:
            user: CustomUser 实例
            start_date: 开始日期 (可选)
            end_date: 结束日期 (可选)
        
        返回:
            HttpResponse 包含 Excel 文件
        """
        check_openpyxl()
        
        # 默认导出最近30天的数据
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        from .models import SleepRecord, SportRecord, Meal, MealItem, BodyMetric
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 1. 睡眠记录工作表
        cls._create_sleep_sheet(wb, user, start_date, end_date, SleepRecord)
        
        # 2. 运动记录工作表
        cls._create_sport_sheet(wb, user, start_date, end_date, SportRecord)
        
        # 3. 饮食记录工作表
        cls._create_diet_sheet(wb, user, start_date, end_date, Meal, MealItem)
        
        # 4. 身体指标工作表
        cls._create_body_metric_sheet(wb, user, start_date, end_date, BodyMetric)
        
        # 删除默认的空白工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        filename = f"健康数据_{user.username}_{start_date}_{end_date}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @classmethod
    def _create_sleep_sheet(cls, wb, user, start_date, end_date, SleepRecord):
        """创建睡眠记录工作表"""
        styles = cls._get_styles()
        ws = wb.create_sheet("睡眠记录")
        
        # 表头
        headers = ["日期", "入睡时间", "起床时间", "睡眠时长(小时)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
            cell.border = styles['thin_border']
        
        # 数据
        records = SleepRecord.objects.filter(
            user=user,
            wakeup_time__date__range=[start_date, end_date]
        ).order_by('wakeup_time')
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.wakeup_time.date().isoformat())
            ws.cell(row=row, column=2, value=record.sleep_time.strftime("%H:%M"))
            ws.cell(row=row, column=3, value=record.wakeup_time.strftime("%H:%M"))
            duration_hours = round(record.duration.total_seconds() / 3600, 2) if record.duration else 0
            ws.cell(row=row, column=4, value=duration_hours)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
    
    @classmethod
    def _create_sport_sheet(cls, wb, user, start_date, end_date, SportRecord):
        """创建运动记录工作表"""
        styles = cls._get_styles()
        ws = wb.create_sheet("运动记录")
        
        headers = ["日期", "运动类型", "时长(分钟)", "消耗热量(大卡)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
        
        records = SportRecord.objects.filter(
            user=user,
            record_date__range=[start_date, end_date]
        ).order_by('record_date')
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.record_date.isoformat())
            ws.cell(row=row, column=2, value=record.sport_type)
            ws.cell(row=row, column=3, value=record.duration_minutes)
            ws.cell(row=row, column=4, value=record.calories_burned)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
    
    @classmethod
    def _create_diet_sheet(cls, wb, user, start_date, end_date, Meal, MealItem):
        """创建饮食记录工作表"""
        styles = cls._get_styles()
        ws = wb.create_sheet("饮食记录")
        
        headers = ["日期", "餐次", "食物名称", "份量(克)", "热量(大卡)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
        
        meals = Meal.objects.filter(
            user=user,
            record_date__range=[start_date, end_date]
        ).prefetch_related('meal_items__food_item').order_by('record_date')
        
        row = 2
        for meal in meals:
            meal_type_display = dict(meal._meta.get_field('meal_type').choices).get(meal.meal_type, meal.meal_type)
            for item in meal.meal_items.all():
                ws.cell(row=row, column=1, value=meal.record_date.isoformat())
                ws.cell(row=row, column=2, value=meal_type_display)
                ws.cell(row=row, column=3, value=item.food_item.name)
                ws.cell(row=row, column=4, value=item.portion)
                ws.cell(row=row, column=5, value=item.calories_calculated)
                row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    @classmethod
    def _create_body_metric_sheet(cls, wb, user, start_date, end_date, BodyMetric):
        """创建身体指标工作表"""
        styles = cls._get_styles()
        ws = wb.create_sheet("身体指标")
        
        headers = ["日期", "体重(kg)", "身高(cm)", "BMI"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
        
        records = BodyMetric.objects.filter(
            user=user,
            record_date__range=[start_date, end_date]
        ).order_by('record_date')
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.record_date.isoformat())
            ws.cell(row=row, column=2, value=record.weight)
            ws.cell(row=row, column=3, value=record.height)
            ws.cell(row=row, column=4, value=record.bmi)
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 12


class ExcelImporter:
    """Excel 导入工具类"""
    
    @classmethod
    def import_food_items(cls, file, update_existing=False):
        """
        从 Excel 文件导入食物数据库
        
        Excel 格式要求:
        - 第一行为表头: 名称, 每100g热量, 蛋白质, 脂肪, 碳水化合物
        - 从第二行开始为数据
        
        参数:
            file: 上传的文件对象
            update_existing: 是否更新已存在的食物
        
        返回:
            dict: {created: int, updated: int, errors: list}
        """
        check_openpyxl()
        from .models import FoodItem
        
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        result = {'created': 0, 'updated': 0, 'errors': []}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                name = str(row[0]).strip() if row[0] else None
                if not name:
                    continue
                    
                calories = float(row[1]) if row[1] else 0
                protein = float(row[2]) if len(row) > 2 and row[2] else None
                fat = float(row[3]) if len(row) > 3 and row[3] else None
                carbs = float(row[4]) if len(row) > 4 and row[4] else None
                
                food, created = FoodItem.objects.get_or_create(
                    name=name,
                    defaults={
                        'calories_per_100g': calories,
                        'protein': protein,
                        'fat': fat,
                        'carbohydrates': carbs
                    }
                )
                
                if created:
                    result['created'] += 1
                elif update_existing:
                    food.calories_per_100g = calories
                    food.protein = protein
                    food.fat = fat
                    food.carbohydrates = carbs
                    food.save()
                    result['updated'] += 1
                    
            except Exception as e:
                result['errors'].append(f"第 {row_num} 行: {str(e)}")
        
        return result


class CSVExporter:
    """CSV 导出工具类 (无需额外依赖)"""
    
    @classmethod
    def export_sleep_csv(cls, user, start_date, end_date):
        """导出睡眠数据为 CSV"""
        from .models import SleepRecord
        
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="sleep_data_{user.username}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['日期', '入睡时间', '起床时间', '睡眠时长(小时)'])
        
        records = SleepRecord.objects.filter(
            user=user,
            wakeup_time__date__range=[start_date, end_date]
        ).order_by('wakeup_time')
        
        for record in records:
            duration_hours = round(record.duration.total_seconds() / 3600, 2) if record.duration else 0
            writer.writerow([
                record.wakeup_time.date().isoformat(),
                record.sleep_time.strftime("%H:%M"),
                record.wakeup_time.strftime("%H:%M"),
                duration_hours
            ])
        
        return response


def export_user_data_json(user):
    """
    导出用户完整数据为 JSON (用于数据备份)
    """
    from .models import SleepRecord, SportRecord, Meal, MealItem, BodyMetric, UserHealthGoal
    
    data = {
        'user': {
            'username': user.username,
            'email': user.email,
            'gender': user.gender,
            'date_of_birth': user.date_of_birth.isoformat() if user.date_of_birth else None
        },
        'export_time': timezone.now().isoformat(),
        'sleep_records': [],
        'sport_records': [],
        'meals': [],
        'body_metrics': []
    }
    
    # 睡眠记录
    for record in SleepRecord.objects.filter(user=user):
        data['sleep_records'].append({
            'sleep_time': record.sleep_time.isoformat(),
            'wakeup_time': record.wakeup_time.isoformat(),
            'duration_seconds': record.duration.total_seconds() if record.duration else 0
        })
    
    # 运动记录
    for record in SportRecord.objects.filter(user=user):
        data['sport_records'].append({
            'sport_type': record.sport_type,
            'duration_minutes': record.duration_minutes,
            'calories_burned': record.calories_burned,
            'record_date': record.record_date.isoformat()
        })
    
    # 饮食记录
    for meal in Meal.objects.filter(user=user).prefetch_related('meal_items__food_item'):
        meal_data = {
            'meal_type': meal.meal_type,
            'record_date': meal.record_date.isoformat(),
            'items': []
        }
        for item in meal.meal_items.all():
            meal_data['items'].append({
                'food_name': item.food_item.name,
                'portion': item.portion,
                'calories': item.calories_calculated
            })
        data['meals'].append(meal_data)
    
    # 身体指标
    for record in BodyMetric.objects.filter(user=user):
        data['body_metrics'].append({
            'weight': record.weight,
            'height': record.height,
            'bmi': record.bmi,
            'record_date': record.record_date.isoformat()
        })
    
    # 返回响应
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="user_data_{user.username}.json"'
    return response
